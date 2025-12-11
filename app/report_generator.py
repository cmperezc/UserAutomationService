"""
Generador de reportes profesionales en PDF y Excel.

Genera reportes detallados del proceso de automatización de usuarios
para auditoría, seguimiento y presentación a directivos.
"""

from pathlib import Path
from datetime import datetime
from typing import Optional
from loguru import logger

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """Genera reportes profesionales del proceso de automatización."""

    # Colores ECR
    COLOR_ECR_BLUE = colors.HexColor("#003366")
    COLOR_SUCCESS = colors.HexColor("#28a745")
    COLOR_ERROR = colors.HexColor("#dc3545")
    COLOR_WARNING = colors.HexColor("#ffc107")
    COLOR_GRAY_BG = colors.HexColor("#f8f9fa")

    # Colores Excel (hex strings sin #)
    EXCEL_HEADER_BG = "003366"
    EXCEL_HEADER_FG = "FFFFFF"
    EXCEL_SUCCESS_BG = "d4edda"
    EXCEL_ERROR_BG = "f8d7da"
    EXCEL_WARNING_BG = "fff3cd"
    EXCEL_BORDER = "dee2e6"

    def __init__(self, output_dir: str = "logs/reportes"):
        """
        Inicializa el generador de reportes.

        Args:
            output_dir: Directorio donde se guardarán los reportes
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    def _create_pdf_styles(self):
        """Crea estilos personalizados para el PDF."""
        styles = getSampleStyleSheet()

        # Título principal
        styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=self.COLOR_ECR_BLUE,
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))

        # Subtítulo
        styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=self.COLOR_ECR_BLUE,
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica'
        ))

        # Encabezado de sección
        styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=self.COLOR_ECR_BLUE,
            spaceAfter=12,
            fontName='Helvetica-Bold'
        ))

        # Texto normal
        styles.add(ParagraphStyle(
            name='CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6
        ))

        return styles

    def generate_pdf_report(self, data: dict) -> str:
        """
        Genera reporte PDF profesional.

        Args:
            data: Diccionario con datos del proceso

        Returns:
            Ruta del archivo PDF generado
        """
        logger.info("Generando reporte PDF...")

        filename = self.output_dir / f"reporte_{self.timestamp}.pdf"
        doc = SimpleDocTemplate(str(filename), pagesize=letter)
        story = []
        styles = self._create_pdf_styles()

        # PÁGINA 1 - PORTADA
        story.append(Spacer(1, 2*inch))
        story.append(Paragraph("Reporte de Automatización de Usuarios", styles['CustomTitle']))
        story.append(Paragraph("Escuela Colombiana de Rehabilitación", styles['CustomSubtitle']))
        story.append(Spacer(1, 0.5*inch))

        # Información del proceso
        info_data = [
            ["Fecha del proceso:", data.get('timestamp', datetime.now().isoformat())],
            ["Archivo procesado:", Path(data.get('excel_file', 'N/A')).name],
        ]
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), self.COLOR_ECR_BLUE),
        ]))
        story.append(info_table)
        story.append(PageBreak())

        # PÁGINA 2 - RESUMEN EJECUTIVO
        story.append(Paragraph("Resumen Ejecutivo", styles['SectionHeader']))
        story.append(Spacer(1, 0.3*inch))

        summary = data.get('summary', {})
        summary_data = [
            ["Métrica", "Cantidad"],
            ["Total usuarios en Excel", str(summary.get('total_in_excel', 0))],
            ["Usuarios nuevos procesados", str(summary.get('new_users', 0))],
            ["Usuarios ya existentes (omitidos)", str(summary.get('existing_users', 0))],
            ["Creados en Office 365", str(summary.get('office365_created', 0))],
            ["Creados en AppConnecto", str(summary.get('appconnecto_created', 0))],
            ["Correos enviados", str(summary.get('emails_sent', 0))],
        ]

        summary_table = Table(summary_data, colWidths=[4*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), self.COLOR_ECR_BLUE),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            # Datos
            ('BACKGROUND', (0, 1), (-1, -1), self.COLOR_GRAY_BG),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1, colors.white),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(summary_table)
        story.append(PageBreak())

        # PÁGINA 3 - DETALLE DE USUARIOS
        story.append(Paragraph("Detalle de Usuarios Procesados", styles['SectionHeader']))
        story.append(Spacer(1, 0.2*inch))

        users = data.get('users', [])
        if users:
            # Limitar a primeros 50 usuarios para no saturar el PDF
            users_to_show = users[:50]

            users_data = [["#", "Nombre", "ID", "Email", "O365", "App", "Email"]]

            for i, user in enumerate(users_to_show, 1):
                name = f"{user.get('full_name', '')} {user.get('full_last_name', '')}"
                id_num = user.get('identification_id', 'N/A')
                email = user.get('institutional_email', 'N/A')

                # Símbolos de status más legibles
                # Office 365: SI (creado), YA EXISTE (ya existía), NO (error al crear)
                if user.get('office365_created'):
                    o365_status = "SI"
                elif user.get('status') == 'existing':
                    o365_status = "YA EXISTE"
                else:
                    o365_status = "NO"

                # AppConnecto: SI (nuevo), YA EXISTE (ya estaba)
                app_status = "SI" if user.get('status') == 'new' else "YA EXISTE"

                # Email: SI (enviado), - (no aplica)
                email_status = "SI" if user.get('password_generated') else "-"

                users_data.append([
                    str(i),
                    name[:25],  # Truncar nombres largos
                    id_num,
                    email[:30],  # Truncar emails largos
                    o365_status,
                    app_status,
                    email_status
                ])

            users_table = Table(
                users_data,
                colWidths=[0.3*inch, 1.6*inch, 0.8*inch, 1.8*inch, 0.7*inch, 0.7*inch, 0.5*inch]
            )
            users_table.setStyle(TableStyle([
                # Encabezado
                ('BACKGROUND', (0, 0), (-1, 0), self.COLOR_ECR_BLUE),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                # Datos - texto general
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (3, -1), 7),  # Columnas de texto más pequeñas
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Número
                # Columnas de status con fuente más pequeña
                ('FONTSIZE', (4, 1), (-1, -1), 6),  # Status más pequeño para que quepa
                ('ALIGN', (4, 1), (-1, -1), 'CENTER'),  # Status centrado
                # Bordes
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.append(users_table)

            if len(users) > 50:
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph(
                    f"<i>Nota: Se muestran los primeros 50 usuarios de {len(users)} totales. "
                    "Consulte el reporte Excel para el listado completo.</i>",
                    styles['CustomNormal']
                ))

        # PÁGINA FINAL - ERRORES
        office365_failed = data.get('office365_results', {}).get('failed', [])
        appconnecto_failed = data.get('appconnecto_results', {}).get('failed', [])
        email_failed = data.get('email_results', {}).get('failed', [])

        if office365_failed or appconnecto_failed or email_failed:
            story.append(PageBreak())
            story.append(Paragraph("Errores y Acciones Requeridas", styles['SectionHeader']))
            story.append(Spacer(1, 0.2*inch))

            if office365_failed:
                story.append(Paragraph("<b>Office 365:</b>", styles['CustomNormal']))
                for user in office365_failed[:10]:  # Primeros 10
                    name = f"{user.get('full_name', '')} {user.get('full_last_name', '')}"
                    error = user.get('creation_error', 'Error desconocido')
                    story.append(Paragraph(f"• {name}: {error}", styles['CustomNormal']))
                story.append(Spacer(1, 0.1*inch))

            if appconnecto_failed:
                story.append(Paragraph("<b>AppConnecto:</b>", styles['CustomNormal']))
                for user in appconnecto_failed[:10]:
                    name = f"{user.get('full_name', '')} {user.get('full_last_name', '')}"
                    story.append(Paragraph(f"• {name}", styles['CustomNormal']))
                story.append(Spacer(1, 0.1*inch))

            if email_failed:
                story.append(Paragraph("<b>Correos:</b>", styles['CustomNormal']))
                for item in email_failed[:10]:
                    name = item.get('name', 'N/A')
                    error = item.get('error', 'Error desconocido')
                    story.append(Paragraph(f"• {name}: {error}", styles['CustomNormal']))

        # Pie de página
        def add_footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.grey)
            canvas.drawString(
                inch,
                0.5*inch,
                f"Generado automáticamente por Sistema de Automatización ECR - {self.timestamp}"
            )
            canvas.drawRightString(
                letter[0] - inch,
                0.5*inch,
                f"Página {canvas.getPageNumber()}"
            )
            canvas.restoreState()

        # Construir PDF
        doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)

        logger.info(f"✅ Reporte PDF generado: {filename}")
        return str(filename)

    def generate_excel_report(self, data: dict) -> str:
        """
        Genera reporte Excel detallado.

        Args:
            data: Mismo diccionario que generate_pdf_report

        Returns:
            Ruta del archivo Excel generado
        """
        logger.info("Generando reporte Excel...")

        filename = self.output_dir / f"reporte_{self.timestamp}.xlsx"
        wb = Workbook()

        # Estilos comunes
        header_font = Font(name='Calibri', size=11, bold=True, color=self.EXCEL_HEADER_FG)
        header_fill = PatternFill(start_color=self.EXCEL_HEADER_BG, end_color=self.EXCEL_HEADER_BG, fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border_side = Side(style='thin', color=self.EXCEL_BORDER)
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        # HOJA 1 - RESUMEN
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"

        # Título
        ws_resumen['A1'] = "Reporte de Automatización de Usuarios"
        ws_resumen['A1'].font = Font(name='Calibri', size=16, bold=True, color=self.EXCEL_HEADER_BG)
        ws_resumen.merge_cells('A1:C1')

        # Información del proceso
        ws_resumen['A3'] = "Fecha del proceso:"
        ws_resumen['B3'] = data.get('timestamp', datetime.now().isoformat())
        ws_resumen['A4'] = "Archivo procesado:"
        ws_resumen['B4'] = Path(data.get('excel_file', 'N/A')).name

        ws_resumen['A3'].font = Font(bold=True)
        ws_resumen['A4'].font = Font(bold=True)

        # Tabla de resumen
        ws_resumen['A6'] = "Métrica"
        ws_resumen['B6'] = "Cantidad"
        ws_resumen['C6'] = "Porcentaje"

        summary = data.get('summary', {})
        total = summary.get('total_in_excel', 1)  # Evitar división por cero

        metrics = [
            ("Total en Excel", summary.get('total_in_excel', 0)),
            ("Nuevos procesados", summary.get('new_users', 0)),
            ("Ya existentes", summary.get('existing_users', 0)),
            ("Creados Office 365", summary.get('office365_created', 0)),
            ("Creados AppConnecto", summary.get('appconnecto_created', 0)),
            ("Correos enviados", summary.get('emails_sent', 0)),
        ]

        row = 7
        for metric, count in metrics:
            ws_resumen[f'A{row}'] = metric
            ws_resumen[f'B{row}'] = count
            percentage = (count / total * 100) if total > 0 else 0
            ws_resumen[f'C{row}'] = f"{percentage:.1f}%"
            row += 1

        # Aplicar estilos a encabezado
        for col in ['A', 'B', 'C']:
            cell = ws_resumen[f'{col}6']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Aplicar bordes a datos
        for r in range(7, row):
            for col in ['A', 'B', 'C']:
                ws_resumen[f'{col}{r}'].border = border
                ws_resumen[f'{col}{r}'].alignment = Alignment(horizontal='left' if col == 'A' else 'center')

        # Ajustar anchos
        ws_resumen.column_dimensions['A'].width = 30
        ws_resumen.column_dimensions['B'].width = 12
        ws_resumen.column_dimensions['C'].width = 12

        # HOJA 2 - USUARIOS
        ws_users = wb.create_sheet("Usuarios")

        headers = [
            "Nombre", "Apellido", "Identificación", "Tipo Doc", "Email Personal",
            "Email Institucional", "Tipo Vinculación", "Programa Académico",
            "Status General", "Office 365", "AppConnecto", "Correo Enviado",
            "Password", "Observaciones"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws_users.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Llenar datos
        users = data.get('users', [])
        for row_num, user in enumerate(users, 2):
            ws_users.cell(row=row_num, column=1, value=user.get('full_name', ''))
            ws_users.cell(row=row_num, column=2, value=user.get('full_last_name', ''))
            ws_users.cell(row=row_num, column=3, value=user.get('identification_id', ''))
            ws_users.cell(row=row_num, column=4, value=user.get('type_document', ''))
            ws_users.cell(row=row_num, column=5, value=user.get('email_personal', ''))
            ws_users.cell(row=row_num, column=6, value=user.get('institutional_email', ''))
            ws_users.cell(row=row_num, column=7, value=user.get('vinculation_type', ''))
            ws_users.cell(row=row_num, column=8, value=user.get('academic_program', ''))

            # Status
            status = user.get('status', 'unknown')
            ws_users.cell(row=row_num, column=9, value=status)

            # Usar texto claro en lugar de símbolos Unicode
            # Office 365: SI (creado), YA EXISTE (ya existía), NO (error al crear)
            if user.get('office365_created'):
                o365 = "SI"
            elif status == 'existing':
                o365 = "YA EXISTE"
            else:
                o365 = "NO"
            ws_users.cell(row=row_num, column=10, value=o365)

            # AppConnecto: SI (nuevo), YA EXISTE (ya estaba)
            app = "SI" if status == 'new' else "YA EXISTE"
            ws_users.cell(row=row_num, column=11, value=app)

            email_sent = "SI" if user.get('password_generated') else "NO"
            ws_users.cell(row=row_num, column=12, value=email_sent)

            # Password (solo si fue generado)
            password = user.get('password_generated', '')
            ws_users.cell(row=row_num, column=13, value=password)

            # Observaciones
            obs = user.get('creation_error', user.get('status_message', ''))
            ws_users.cell(row=row_num, column=14, value=obs)

            # Aplicar bordes
            for col in range(1, 15):
                ws_users.cell(row=row_num, column=col).border = border
                ws_users.cell(row=row_num, column=col).alignment = Alignment(vertical='center')

            # Color de fondo según status
            if user.get('office365_created'):
                fill = PatternFill(start_color=self.EXCEL_SUCCESS_BG, end_color=self.EXCEL_SUCCESS_BG, fill_type='solid')
            elif user.get('creation_error'):
                fill = PatternFill(start_color=self.EXCEL_ERROR_BG, end_color=self.EXCEL_ERROR_BG, fill_type='solid')
            else:
                fill = PatternFill(start_color=self.EXCEL_WARNING_BG, end_color=self.EXCEL_WARNING_BG, fill_type='solid')

            for col in range(1, 15):
                ws_users.cell(row=row_num, column=col).fill = fill

        # Ajustar anchos
        column_widths = [15, 15, 12, 10, 25, 25, 15, 30, 12, 10, 12, 12, 15, 40]
        for i, width in enumerate(column_widths, 1):
            ws_users.column_dimensions[get_column_letter(i)].width = width

        # Filtros
        ws_users.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(users) + 1}"

        # Congelar primera fila
        ws_users.freeze_panes = "A2"

        # HOJA 3 - ERRORES
        ws_errors = wb.create_sheet("Errores")

        error_headers = ["Usuario", "Plataforma", "Error", "Fecha", "Acción Requerida"]
        for col_num, header in enumerate(error_headers, 1):
            cell = ws_errors.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Recopilar errores
        error_row = 2
        office365_failed = data.get('office365_results', {}).get('failed', [])
        for user in office365_failed:
            name = f"{user.get('full_name', '')} {user.get('full_last_name', '')}"
            ws_errors.cell(row=error_row, column=1, value=name)
            ws_errors.cell(row=error_row, column=2, value="Office 365")
            ws_errors.cell(row=error_row, column=3, value=user.get('creation_error', 'Error desconocido'))
            ws_errors.cell(row=error_row, column=4, value=self.timestamp)
            ws_errors.cell(row=error_row, column=5, value="Revisar y reintentar creación manual")

            for col in range(1, 6):
                ws_errors.cell(row=error_row, column=col).border = border

            error_row += 1

        appconnecto_failed = data.get('appconnecto_results', {}).get('failed', [])
        for user in appconnecto_failed:
            name = f"{user.get('full_name', '')} {user.get('full_last_name', '')}"
            ws_errors.cell(row=error_row, column=1, value=name)
            ws_errors.cell(row=error_row, column=2, value="AppConnecto")
            ws_errors.cell(row=error_row, column=3, value="Error en creación")
            ws_errors.cell(row=error_row, column=4, value=self.timestamp)
            ws_errors.cell(row=error_row, column=5, value="Crear manualmente en AppConnecto")

            for col in range(1, 6):
                ws_errors.cell(row=error_row, column=col).border = border

            error_row += 1

        email_failed = data.get('email_results', {}).get('failed', [])
        for item in email_failed:
            ws_errors.cell(row=error_row, column=1, value=item.get('name', 'N/A'))
            ws_errors.cell(row=error_row, column=2, value="Email")
            ws_errors.cell(row=error_row, column=3, value=item.get('error', 'Error desconocido'))
            ws_errors.cell(row=error_row, column=4, value=self.timestamp)
            ws_errors.cell(row=error_row, column=5, value="Reenviar credenciales manualmente")

            for col in range(1, 6):
                ws_errors.cell(row=error_row, column=col).border = border

            error_row += 1

        # Ajustar anchos
        ws_errors.column_dimensions['A'].width = 30
        ws_errors.column_dimensions['B'].width = 15
        ws_errors.column_dimensions['C'].width = 40
        ws_errors.column_dimensions['D'].width = 20
        ws_errors.column_dimensions['E'].width = 35

        # Guardar
        wb.save(filename)

        logger.info(f"✅ Reporte Excel generado: {filename}")
        return str(filename)

    def generate_both(self, data: dict) -> dict:
        """
        Genera ambos reportes (PDF y Excel).

        Args:
            data: Diccionario con datos del proceso

        Returns:
            Dict con rutas: {"pdf": "...", "excel": "..."}
        """
        logger.info("Generando reportes PDF y Excel...")

        pdf_path = self.generate_pdf_report(data)
        excel_path = self.generate_excel_report(data)

        return {
            "pdf": pdf_path,
            "excel": excel_path
        }
