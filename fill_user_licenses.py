"""
Script para rellenar información de correo electrónico y licencias de estudiantes
consultando Microsoft Graph API por nombre.
"""

import asyncio
import openpyxl
from loguru import logger
from app.graph_api import get_graph_client
from unidecode import unidecode


def normalize_name(name: str) -> str:
    """
    Normaliza un nombre para comparación:
    - Convierte a minúsculas
    - Elimina acentos
    - Elimina espacios extras
    """
    if not name:
        return ""
    name = unidecode(name.strip().lower())
    # Eliminar espacios múltiples
    name = " ".join(name.split())
    return name


def names_match(name1: str, name2: str) -> bool:
    """
    Compara dos nombres teniendo en cuenta diferentes formatos.

    name1: Formato Excel "CASTRO GIL SHARITK NICOL"
    name2: Formato Office365 "Sharitk Nicol Castro Gil"

    Returns:
        True si los nombres coinciden (tienen las mismas palabras)
    """
    # Normalizar ambos nombres
    norm1 = normalize_name(name1)
    norm2 = normalize_name(name2)

    # Separar en palabras
    words1 = set(norm1.split())
    words2 = set(norm2.split())

    # Si no hay palabras, no hay coincidencia
    if not words1 or not words2:
        return False

    # Calcular intersección de palabras
    common_words = words1.intersection(words2)

    # Si al menos el 70% de las palabras coinciden, considerarlo un match
    # Esto permite cierta flexibilidad en caso de nombres con más partes
    min_len = min(len(words1), len(words2))
    if min_len == 0:
        return False

    match_threshold = 0.7
    match_ratio = len(common_words) / min_len

    return match_ratio >= match_threshold


async def search_user_by_name(client, student_name: str, domain: str = "ecr.edu.co"):
    """
    Busca un usuario en Microsoft Graph API por su nombre.

    Args:
        client: Cliente de Graph API
        student_name: Nombre del estudiante a buscar
        domain: Dominio del email

    Returns:
        Dict con email y licencias del usuario, o None si no se encuentra
    """
    logger.info(f"Buscando usuario: {student_name}")

    token = client.get_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "ConsistencyLevel": "eventual"
    }

    # Normalizar nombre para comparación
    normalized_search = normalize_name(student_name)

    # Buscar TODOS los usuarios con paginación
    import aiohttp
    url = f"https://graph.microsoft.com/v1.0/users?$select=displayName,mail,userPrincipalName,assignedLicenses&$top=999"

    all_users = []

    async with aiohttp.ClientSession() as session:
        # Paginación: seguir @odata.nextLink hasta obtener todos
        while url:
            async with session.get(url, headers=headers) as response:
                if response.status == 200:
                    data = await response.json()
                    users = data.get('value', [])
                    all_users.extend(users)

                    # Siguiente página
                    url = data.get('@odata.nextLink')
                else:
                    error_text = await response.text()
                    logger.error(f"❌ Error obteniendo usuarios: {error_text}")
                    break

        logger.debug(f"Total de usuarios obtenidos: {len(all_users)}")

        # Buscar coincidencia por nombre
        for user in all_users:
            display_name = user.get('displayName', '')

            email = user.get('mail') or user.get('userPrincipalName')

            # Filtrar solo usuarios del dominio correcto
            if not email or not email.lower().endswith(f"@{domain}"):
                continue

            # Comparar nombres usando lógica mejorada
            if names_match(student_name, display_name):
                licenses = user.get('assignedLicenses', [])

                # Obtener SKUs de licencias
                license_names = await get_license_names(client, licenses)

                logger.info(f"✅ Usuario encontrado: {display_name} - {email}")
                logger.info(f"   Licencias: {', '.join(license_names) if license_names else 'Sin licencias'}")

                return {
                    'email': email,
                    'licenses': ', '.join(license_names) if license_names else 'Sin licencias',
                    'display_name': display_name
                }

        logger.warning(f"⚠️ No se encontró usuario con nombre: {student_name}")
        logger.debug(f"   Búsqueda normalizada: '{normalized_search}'")
        return None


async def get_license_names(client, assigned_licenses: list) -> list[str]:
    """
    Convierte SKU IDs de licencias a nombres legibles.

    Args:
        client: Cliente de Graph API
        assigned_licenses: Lista de licencias asignadas con skuId

    Returns:
        Lista de nombres de licencias
    """
    if not assigned_licenses:
        return []

    token = client.get_token()
    headers = {"Authorization": f"Bearer {token}"}

    license_names = []

    import aiohttp
    async with aiohttp.ClientSession() as session:
        # Obtener información de SKUs suscritas
        async with session.get(
            "https://graph.microsoft.com/v1.0/subscribedSkus",
            headers=headers
        ) as response:
            if response.status == 200:
                data = await response.json()
                skus = data.get('value', [])

                # Crear mapeo de skuId -> nombre
                sku_map = {sku['skuId']: sku['skuPartNumber'] for sku in skus}

                # Obtener nombres de licencias asignadas
                for license in assigned_licenses:
                    sku_id = license.get('skuId')
                    if sku_id in sku_map:
                        license_names.append(sku_map[sku_id])
            else:
                logger.warning("No se pudieron obtener nombres de licencias")

    return license_names


async def fill_licenses_from_excel(excel_path: str, output_path: str = None):
    """
    Rellena campos de correo y licencia en Excel consultando Microsoft Graph API.

    Args:
        excel_path: Ruta al archivo Excel de entrada
        output_path: Ruta al archivo Excel de salida (opcional, sobreescribe si no se especifica)
    """
    if output_path is None:
        output_path = excel_path

    logger.info(f"Procesando archivo: {excel_path}")

    # Cargar Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Obtener cliente de Graph API
    client = get_graph_client()

    # Identificar columnas (asumiendo primera fila es encabezado)
    headers = [cell.value for cell in ws[1]]
    logger.info(f"Columnas encontradas: {headers}")

    # Buscar índices de columnas
    col_nombre = None
    col_email = None
    col_licencia = None

    for idx, header in enumerate(headers, start=1):
        header_lower = str(header).lower().strip() if header else ""
        if 'nombre' in header_lower and 'estudiante' in header_lower:
            col_nombre = idx
        elif 'correo' in header_lower or 'email' in header_lower:
            col_email = idx
        elif 'licencia' in header_lower:
            col_licencia = idx

    if not col_nombre:
        logger.error("❌ No se encontró columna de nombre de estudiante")
        return
    if not col_email:
        logger.error("❌ No se encontró columna de correo electrónico")
        return
    if not col_licencia:
        logger.error("❌ No se encontró columna de licencia")
        return

    logger.info(f"Columnas: Nombre={col_nombre}, Email={col_email}, Licencia={col_licencia}")

    # Procesar filas (desde la fila 2, asumiendo fila 1 es encabezado)
    total_rows = ws.max_row
    processed = 0
    found = 0
    not_found = 0

    for row_num in range(2, total_rows + 1):
        student_name = ws.cell(row=row_num, column=col_nombre).value

        if not student_name:
            logger.debug(f"Fila {row_num}: Sin nombre, omitiendo")
            continue

        processed += 1
        logger.info(f"\n--- Procesando {processed}/{total_rows - 1}: {student_name} ---")

        # Buscar usuario en Graph API
        user_info = await search_user_by_name(client, student_name)

        if user_info:
            # Rellenar campos
            ws.cell(row=row_num, column=col_email, value=user_info['email'])
            ws.cell(row=row_num, column=col_licencia, value=user_info['licenses'])
            found += 1
            logger.info(f"✅ Fila {row_num}: Datos completados")
        else:
            # Marcar como no encontrado
            ws.cell(row=row_num, column=col_email, value="NO ENCONTRADO")
            ws.cell(row=row_num, column=col_licencia, value="NO ENCONTRADO")
            not_found += 1
            logger.warning(f"⚠️ Fila {row_num}: Usuario no encontrado")

    # Guardar cambios
    wb.save(output_path)
    logger.info(f"\n{'='*60}")
    logger.info(f"✅ Archivo guardado: {output_path}")
    logger.info(f"📊 Resumen:")
    logger.info(f"   - Total procesados: {processed}")
    logger.info(f"   - Encontrados: {found}")
    logger.info(f"   - No encontrados: {not_found}")
    logger.info(f"{'='*60}")


async def main():
    """Función principal"""
    excel_file = "Licencia.xlsx"
    output_file = "Licencia_Actualizado.xlsx"

    try:
        await fill_licenses_from_excel(excel_file, output_file)
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        raise


if __name__ == "__main__":
    # Configurar logging
    logger.add(
        "logs/fill_licenses.log",
        rotation="10 MB",
        level="DEBUG",
        format="{time:YYYY-MM-DD HH:mm:ss} | {level: <8} | {message}"
    )

    asyncio.run(main())
