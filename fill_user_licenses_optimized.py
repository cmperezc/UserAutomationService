"""
Script OPTIMIZADO para rellenar información de correo electrónico y licencias de estudiantes
consultando Microsoft Graph API por nombre.

OPTIMIZACIÓN: Carga todos los usuarios UNA SOLA VEZ al inicio.
"""

import asyncio
import openpyxl
from loguru import logger
from app.graph_api import get_graph_client
from unidecode import unidecode


# Mapeo de códigos técnicos a nombres legibles de licencias
LICENSE_NAMES = {
    'STANDARDWOFFPACK_STUDENT': 'Office 365 A1 para estudiantes',
    'STANDARDWOFFPACK_FACULTY': 'Office 365 A1 para profesores',
    'M365EDU_A5_STUUSEBNFT': 'Microsoft 365 A5 para estudiantes',
    'M365EDU_A5_FACULTY': 'Microsoft 365 A5 para profesores',
    'ENTERPRISEPREMIUM_STUUSEBNFT': 'Office 365 A3 para estudiantes',
    'FLOW_FREE': 'Microsoft Power Automate Free',
    'POWER_BI_STANDARD': 'Power BI Standard',
    'STREAM': 'Microsoft Stream',
    'WINDOWS_STORE': 'Windows Store',
    'CCIBOTS_PRIVPREV_VIRAL': 'Power Virtual Agents',
    'Microsoft_365_Copilot_EDU': 'Microsoft 365 Copilot Educación',
    'PROJECTPROFESSIONAL_FACULTY': 'Project Professional para profesores',
    'Power_Pages_vTrial_for_Makers': 'Power Pages Trial',
    'Teams_Premium_(for_Departments)': 'Teams Premium',
    'PHONESYSTEM_VIRTUALUSER_FACULTY': 'Phone System Virtual User',
    'VISIOCLIENT_FACULTY': 'Visio para profesores',
    'POWERAPPS_DEV': 'Power Apps Developer',
}


def normalize_name(name: str) -> str:
    """
    Normaliza un nombre para comparación:
    - Convierte a minúsculas
    - Elimina acentos
    - Elimina espacios extras
    """
    if not name:
        return ""
    name = unidecode(name.strip().lower())
    # Eliminar espacios múltiples
    name = " ".join(name.split())
    return name


def names_match(name1: str, name2: str) -> bool:
    """
    Compara dos nombres teniendo en cuenta diferentes formatos.

    name1: Formato Excel "CASTRO GIL SHARITK NICOL"
    name2: Formato Office365 "Sharitk Nicol Castro Gil"

    Returns:
        True si los nombres coinciden (tienen las mismas palabras)
    """
    # Normalizar ambos nombres
    norm1 = normalize_name(name1)
    norm2 = normalize_name(name2)

    # Separar en palabras
    words1 = set(norm1.split())
    words2 = set(norm2.split())

    # Si no hay palabras, no hay coincidencia
    if not words1 or not words2:
        return False

    # Calcular intersección de palabras
    common_words = words1.intersection(words2)

    # Si al menos el 70% de las palabras coinciden, considerarlo un match
    # Esto permite cierta flexibilidad en caso de nombres con más partes
    min_len = min(len(words1), len(words2))
    if min_len == 0:
        return False

    match_threshold = 0.7
    match_ratio = len(common_words) / min_len

    return match_ratio >= match_threshold


async def load_all_users(client, domain: str = "ecr.edu.co"):
    """
    Carga TODOS los usuarios del dominio UNA SOLA VEZ.

    Returns:
        Lista de usuarios con email, displayName y licencias
    """
    logger.info(f"🔄 Cargando todos los usuarios del dominio @{domain}...")

    token = client.get_token()
    headers = {"Authorization": f"Bearer {token}"}

    all_users = []
    url = f"https://graph.microsoft.com/v1.0/users?$select=displayName,mail,userPrincipalName,assignedLicenses&$top=999"

    import aiohttp
    async with aiohttp.ClientSession() as session:
        # Paginación: seguir @odata.nextLink hasta obtener todos
        while url:
            async with session.get(url, headers=headers) as response:
                if response.status == 200:
                    data = await response.json()
                    users = data.get('value', [])

                    # Filtrar solo usuarios del dominio correcto
                    for user in users:
                        email = user.get('mail') or user.get('userPrincipalName')
                        if email and email.lower().endswith(f"@{domain}"):
                            all_users.append(user)

                    # Siguiente página
                    url = data.get('@odata.nextLink')
                else:
                    error_text = await response.text()
                    logger.error(f"❌ Error obteniendo usuarios: {error_text}")
                    break

    logger.info(f"✅ Total de usuarios cargados: {len(all_users)}")
    return all_users


async def get_license_names(client, assigned_licenses: list) -> list[str]:
    """
    Convierte SKU IDs de licencias a nombres legibles.
    """
    if not assigned_licenses:
        return []

    token = client.get_token()
    headers = {"Authorization": f"Bearer {token}"}

    license_names = []

    import aiohttp
    async with aiohttp.ClientSession() as session:
        async with session.get(
            "https://graph.microsoft.com/v1.0/subscribedSkus",
            headers=headers
        ) as response:
            if response.status == 200:
                data = await response.json()
                skus = data.get('value', [])

                # Crear mapeo de skuId -> nombre
                sku_map = {sku['skuId']: sku['skuPartNumber'] for sku in skus}

                # Obtener nombres de licencias asignadas
                for license in assigned_licenses:
                    sku_id = license.get('skuId')
                    if sku_id in sku_map:
                        license_names.append(sku_map[sku_id])
            else:
                logger.warning("No se pudieron obtener nombres de licencias")

    return license_names


def search_user_in_list(student_name: str, all_users: list) -> dict | None:
    """
    Busca un usuario en la lista precargada por su nombre.

    Args:
        student_name: Nombre del estudiante a buscar
        all_users: Lista de usuarios precargada

    Returns:
        Dict con displayName, email y assignedLicenses del usuario, o None
    """
    # Buscar coincidencia por nombre
    for user in all_users:
        display_name = user.get('displayName', '')

        # Comparar nombres usando lógica mejorada
        if names_match(student_name, display_name):
            return user

    return None


async def fill_licenses_from_excel(excel_path: str, output_path: str = None):
    """
    Rellena campos de correo y licencia en Excel consultando Microsoft Graph API.

    OPTIMIZADO: Carga todos los usuarios UNA SOLA VEZ.

    Args:
        excel_path: Ruta al archivo Excel de entrada
        output_path: Ruta al archivo Excel de salida (opcional, sobreescribe si no se especifica)
    """
    if output_path is None:
        output_path = excel_path.replace('.xlsx', '_Actualizado.xlsx')

    logger.info(f"📂 Procesando archivo: {excel_path}")

    # Cargar Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Obtener cliente de Graph API
    client = get_graph_client()

    # 🚀 OPTIMIZACIÓN: Cargar TODOS los usuarios UNA SOLA VEZ
    all_users = await load_all_users(client)

    # Obtener mapeo de licencias UNA SOLA VEZ
    logger.info("🔄 Cargando información de licencias...")
    token = client.get_token()
    import aiohttp
    async with aiohttp.ClientSession() as session:
        async with session.get(
            "https://graph.microsoft.com/v1.0/subscribedSkus",
            headers={"Authorization": f"Bearer {token}"}
        ) as response:
            if response.status == 200:
                data = await response.json()
                skus = data.get('value', [])
                sku_map = {sku['skuId']: sku['skuPartNumber'] for sku in skus}
                logger.info(f"✅ {len(sku_map)} tipos de licencias disponibles")
            else:
                sku_map = {}
                logger.warning("No se pudieron cargar licencias")

    # Identificar columnas (asumiendo primera fila es encabezado)
    headers = [cell.value for cell in ws[1]]
    logger.info(f"📋 Columnas encontradas: {headers}")

    # Buscar índices de columnas
    col_nombre = None
    col_email = None
    col_licencia = None

    for idx, header in enumerate(headers, start=1):
        header_lower = str(header).lower().strip() if header else ""
        if 'nombre' in header_lower and 'estudiante' in header_lower:
            col_nombre = idx
        elif 'correo' in header_lower or 'email' in header_lower:
            col_email = idx
        elif 'licencia' in header_lower:
            col_licencia = idx

    if not col_nombre:
        logger.error("❌ No se encontró columna de nombre de estudiante")
        return
    if not col_email:
        logger.error("❌ No se encontró columna de correo electrónico")
        return
    if not col_licencia:
        logger.error("❌ No se encontró columna de licencia")
        return

    logger.info(f"📍 Columnas: Nombre={col_nombre}, Email={col_email}, Licencia={col_licencia}")

    # Procesar filas (desde la fila 2, asumiendo fila 1 es encabezado)
    total_rows = ws.max_row
    processed = 0
    found = 0
    not_found = 0

    for row_num in range(2, total_rows + 1):
        student_name = ws.cell(row=row_num, column=col_nombre).value

        if not student_name:
            logger.debug(f"Fila {row_num}: Sin nombre, omitiendo")
            continue

        processed += 1
        logger.info(f"\n{'='*60}")
        logger.info(f"[{processed}/{total_rows - 1}] Buscando: {student_name}")

        # Buscar usuario en la lista precargada
        user = search_user_in_list(student_name, all_users)

        if user:
            display_name = user.get('displayName', '')
            email = user.get('mail') or user.get('userPrincipalName')
            licenses = user.get('assignedLicenses', [])

            # Convertir licencias a nombres legibles
            license_codes = [sku_map.get(lic['skuId'], lic['skuId']) for lic in licenses]
            license_names = [LICENSE_NAMES.get(code, code) for code in license_codes]
            license_str = ', '.join(license_names) if license_names else 'Sin licencias'

            # Rellenar campos
            ws.cell(row=row_num, column=col_email, value=email)
            ws.cell(row=row_num, column=col_licencia, value=license_str)

            found += 1
            logger.info(f"✅ Encontrado: {display_name}")
            logger.info(f"   📧 Email: {email}")
            logger.info(f"   📝 Licencias: {license_str}")
        else:
            # Marcar como no encontrado
            ws.cell(row=row_num, column=col_email, value="NO ENCONTRADO")
            ws.cell(row=row_num, column=col_licencia, value="NO ENCONTRADO")
            not_found += 1
            logger.warning(f"⚠️ NO encontrado")

    # Guardar cambios
    wb.save(output_path)
    logger.info(f"\n{'='*60}")
    logger.info(f"✅ Archivo guardado: {output_path}")
    logger.info(f"\n📊 RESUMEN FINAL:")
    logger.info(f"{'='*60}")
    logger.info(f"   📝 Total procesados: {processed}")
    logger.info(f"   ✅ Encontrados: {found} ({found/processed*100:.1f}%)" if processed > 0 else "   ✅ Encontrados: 0")
    logger.info(f"   ❌ No encontrados: {not_found} ({not_found/processed*100:.1f}%)" if processed > 0 else "   ❌ No encontrados: 0")
    logger.info(f"{'='*60}")


async def main():
    """Función principal"""
    excel_file = "Licencia.xlsx"

    # Generar nombre con timestamp para evitar conflictos
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"Licencia_Completo_{timestamp}.xlsx"

    try:
        await fill_licenses_from_excel(excel_file, output_file)
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        raise


if __name__ == "__main__":
    # Configurar logging
    logger.add(
        "logs/fill_licenses.log",
        rotation="10 MB",
        level="DEBUG",
        format="{time:YYYY-MM-DD HH:mm:ss} | {level: <8} | {message}"
    )

    asyncio.run(main())
