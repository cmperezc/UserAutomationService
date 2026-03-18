"""
Script para verificar que los usuarios están en el grupo de licencia correcto
según su tipo de trabajador en el Excel activos.xlsx.

Reglas:
  - PASANTE SENA, ADMINISTRATIVO, PRACTICANTE -> "Administrativos Licencia A5"
  - DOCENTE -> "Docentes Licencia A5"
"""

import asyncio
import openpyxl
import aiohttp
from pathlib import Path
from datetime import datetime
from app.graph_api import get_graph_client
from loguru import logger

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_CENTER


# Mapeo de tipo de trabajador -> grupo esperado
GRUPO_ESPERADO = {
    "PASANTE SENA": "Administrativos Licencia A5",
    "ADMINISTRATIVO": "Administrativos Licencia A5",
    "PRACTICANTE": "Administrativos Licencia A5",
    "DOCENTE": "Docentes Licencia A5",
}

GRUPOS_LICENCIA = [
    "Administrativos Licencia A5",
    "Docentes Licencia A5",
]

GRUPO_RETIRADOS = "Administrativos y Docentes Retirados Licencia A1"


async def get_group_members(client, group_name: str) -> dict[str, str]:
    """
    Obtiene todos los miembros de un grupo.

    Returns:
        Dict de email (minúsculas) -> displayName
    """
    group_id = await client.get_group_id(group_name)
    if not group_id:
        logger.error(f"No se encontró el grupo: {group_name}")
        return {}

    token = client.get_token()
    headers = {"Authorization": f"Bearer {token}"}

    members = {}
    url = f"https://graph.microsoft.com/v1.0/groups/{group_id}/members?$select=displayName,mail,userPrincipalName&$top=999"

    async with aiohttp.ClientSession() as session:
        while url:
            async with session.get(url, headers=headers) as response:
                if response.status == 200:
                    data = await response.json()
                    for member in data.get('value', []):
                        email = member.get('mail') or member.get('userPrincipalName')
                        if email:
                            members[email.lower()] = member.get('displayName', '')
                    url = data.get('@odata.nextLink')
                else:
                    error_text = await response.text()
                    logger.error(f"Error obteniendo miembros de '{group_name}': {error_text}")
                    break

    logger.info(f"Grupo '{group_name}': {len(members)} miembros")
    return members


async def check_group_assignments(excel_path: str = "activos.xlsx"):
    """
    Lee el Excel y verifica que cada usuario esté en el grupo correcto.
    """
    client = get_graph_client()

    # 1. Cargar miembros de cada grupo (email -> displayName)
    logger.info("Cargando miembros de los grupos de licencia...")
    grupo_miembros = {}
    for grupo in GRUPOS_LICENCIA:
        grupo_miembros[grupo] = await get_group_members(client, grupo)

    # 2. Leer Excel
    logger.info(f"Leyendo archivo: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    col_tipo = None
    col_nombre = None
    col_email = None

    for idx, header in enumerate(headers):
        header_upper = header.upper()
        if "TIPO DE TRABAJADOR" in header_upper:
            col_tipo = idx
        elif "NOMBRE COMPLETO" in header_upper:
            col_nombre = idx
        elif "EMAIL_INSTITUCIONAL" in header_upper or "EMAIL" in header_upper:
            col_email = idx

    if col_tipo is None or col_nombre is None or col_email is None:
        logger.error(f"No se encontraron las columnas requeridas. Columnas: {headers}")
        return

    # 3. Verificar cada usuario
    total_filas = ws.max_row - 1  # Sin encabezado
    correctos = []
    incorrectos = []
    sin_grupo = []
    tipo_invalido = []
    omitidos = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        tipo_trabajador = str(row[col_tipo]).strip().upper() if row[col_tipo] else ""
        nombre = str(row[col_nombre]).strip() if row[col_nombre] else ""
        email = str(row[col_email]).strip().lower() if row[col_email] else ""

        if not email or not tipo_trabajador:
            omitidos.append({
                "nombre": nombre or "(vacio)",
                "email": email or "(vacio)",
                "tipo": tipo_trabajador or "(vacio)",
            })
            continue

        # Validar tipo de trabajador
        grupo_esperado = GRUPO_ESPERADO.get(tipo_trabajador)
        if not grupo_esperado:
            tipo_invalido.append({
                "nombre": nombre,
                "email": email,
                "tipo": tipo_trabajador,
            })
            continue

        # Determinar en qué grupo(s) está actualmente
        grupos_actuales = []
        for grupo, miembros in grupo_miembros.items():
            if email in miembros:
                grupos_actuales.append(grupo)

        if not grupos_actuales:
            sin_grupo.append({
                "nombre": nombre,
                "email": email,
                "tipo": tipo_trabajador,
                "grupo_esperado": grupo_esperado,
            })
        elif grupo_esperado in grupos_actuales and len(grupos_actuales) == 1:
            correctos.append({
                "nombre": nombre,
                "email": email,
                "tipo": tipo_trabajador,
                "grupo_actual": grupos_actuales[0],
            })
        else:
            incorrectos.append({
                "nombre": nombre,
                "email": email,
                "tipo": tipo_trabajador,
                "grupo_esperado": grupo_esperado,
                "grupos_actuales": grupos_actuales,  # lista original, sin join
            })

    # 4. Detectar sobrantes: usuarios en los grupos que NO están en el Excel
    emails_excel = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        email_val = str(row[col_email]).strip().lower() if row[col_email] else ""
        if email_val:
            emails_excel.add(email_val)

    sobrantes = []
    for grupo, miembros in grupo_miembros.items():
        for email, display_name in miembros.items():
            if email not in emails_excel:
                sobrantes.append({
                    "nombre": display_name or "(sin nombre)",
                    "email": email,
                    "grupo": grupo,
                })

    # 5. Imprimir reporte
    print("\n" + "=" * 90)
    print("REPORTE DE VERIFICACION DE GRUPOS DE LICENCIA")
    print("=" * 90)

    total_procesados = len(correctos) + len(incorrectos) + len(sin_grupo) + len(tipo_invalido)
    print(f"\nTotal filas en Excel:  {total_filas}")
    print(f"Total procesados:      {total_procesados}")
    print(f"Omitidos (sin email/tipo): {len(omitidos)}")
    print(f"  Correctos:           {len(correctos)}")
    print(f"  En grupo incorrecto: {len(incorrectos)}")
    print(f"  Sin grupo asignado:  {len(sin_grupo)}")
    print(f"  Tipo invalido:       {len(tipo_invalido)}")
    print(f"  Sobrantes (en grupo pero no en Excel): {len(sobrantes)}")

    if incorrectos:
        print("\n" + "-" * 90)
        print("USUARIOS EN GRUPO INCORRECTO")
        print("-" * 90)
        print(f"{'NOMBRE':<40} {'TIPO':<18} {'ACTUAL':<30} {'ESPERADO'}")
        print("-" * 90)
        for u in incorrectos:
            print(f"{u['nombre']:<40} {u['tipo']:<18} {', '.join(u['grupos_actuales']):<30} {u['grupo_esperado']}")

    if sin_grupo:
        print("\n" + "-" * 90)
        print("USUARIOS SIN GRUPO DE LICENCIA")
        print("-" * 90)
        print(f"{'NOMBRE':<40} {'EMAIL':<35} {'TIPO':<18} {'DEBERIA ESTAR EN'}")
        print("-" * 90)
        for u in sin_grupo:
            print(f"{u['nombre']:<40} {u['email']:<35} {u['tipo']:<18} {u['grupo_esperado']}")

    if tipo_invalido:
        print("\n" + "-" * 90)
        print("USUARIOS CON TIPO DE TRABAJADOR NO RECONOCIDO")
        print("-" * 90)
        for u in tipo_invalido:
            print(f"  {u['nombre']} | {u['email']} | Tipo: '{u['tipo']}'")

    if sobrantes:
        print("\n" + "-" * 90)
        print(f"SOBRANTES: EN GRUPO PERO NO EN EL EXCEL (se moverán a '{GRUPO_RETIRADOS}')")
        print("-" * 90)
        print(f"{'NOMBRE':<40} {'EMAIL':<35} {'GRUPO ACTUAL'}")
        print("-" * 90)
        for u in sobrantes:
            print(f"{u['nombre']:<40} {u['email']:<35} {u['grupo']}")

    if omitidos:
        print("\n" + "-" * 90)
        print("FILAS OMITIDAS (sin email o sin tipo de trabajador)")
        print("-" * 90)
        for u in omitidos:
            print(f"  Nombre: {u['nombre']} | Email: {u['email']} | Tipo: {u['tipo']}")

    print("\n" + "=" * 90)

    # 6. Generar PDF
    pdf_path = generate_pdf_report(total_filas, total_procesados, correctos, incorrectos, sin_grupo, tipo_invalido, omitidos, sobrantes)
    print(f"\nReporte PDF generado: {pdf_path}")

    # 7. Acciones interactivas
    se_aplico_algo = await aplicar_cambios_interactivo(client, incorrectos, sin_grupo, sobrantes)

    # 8. Si se aplicaron cambios, re-consultar y generar PDF post-corrección
    if se_aplico_algo:
        print("\nRe-consultando el estado actual para generar reporte post-corrección...")
        grupo_miembros_post = {}
        for grupo in GRUPOS_LICENCIA:
            grupo_miembros_post[grupo] = await get_group_members(client, grupo)

        correctos_post, incorrectos_post, sin_grupo_post, tipo_invalido_post, omitidos_post = [], [], [], [], []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            tipo_trabajador = str(row[col_tipo]).strip().upper() if row[col_tipo] else ""
            nombre = str(row[col_nombre]).strip() if row[col_nombre] else ""
            email = str(row[col_email]).strip().lower() if row[col_email] else ""

            if not email or not tipo_trabajador:
                omitidos_post.append({"nombre": nombre or "(vacio)", "email": email or "(vacio)", "tipo": tipo_trabajador or "(vacio)"})
                continue

            grupo_esperado = GRUPO_ESPERADO.get(tipo_trabajador)
            if not grupo_esperado:
                tipo_invalido_post.append({"nombre": nombre, "email": email, "tipo": tipo_trabajador})
                continue

            grupos_actuales_post = [g for g, miembros in grupo_miembros_post.items() if email in miembros]
            if not grupos_actuales_post:
                sin_grupo_post.append({"nombre": nombre, "email": email, "tipo": tipo_trabajador, "grupo_esperado": grupo_esperado})
            elif grupo_esperado in grupos_actuales_post and len(grupos_actuales_post) == 1:
                correctos_post.append({"nombre": nombre, "email": email, "tipo": tipo_trabajador, "grupo_actual": grupos_actuales_post[0]})
            else:
                incorrectos_post.append({"nombre": nombre, "email": email, "tipo": tipo_trabajador, "grupo_esperado": grupo_esperado, "grupos_actuales": grupos_actuales_post})

        sobrantes_post = []
        for grupo, miembros in grupo_miembros_post.items():
            for email, display_name in miembros.items():
                if email not in emails_excel:
                    sobrantes_post.append({"nombre": display_name or "(sin nombre)", "email": email, "grupo": grupo})

        total_procesados_post = len(correctos_post) + len(incorrectos_post) + len(sin_grupo_post) + len(tipo_invalido_post)
        pdf_post = generate_pdf_report(
            total_filas, total_procesados_post,
            correctos_post, incorrectos_post, sin_grupo_post,
            tipo_invalido_post, omitidos_post, sobrantes_post,
            es_post_correccion=True,
        )
        print(f"Reporte PDF post-corrección generado: {pdf_post}")


def generate_pdf_report(
    total_filas: int,
    total_procesados: int,
    correctos: list,
    incorrectos: list,
    sin_grupo: list,
    tipo_invalido: list,
    omitidos: list,
    sobrantes: list,
    es_post_correccion: bool = False,
) -> str:
    """Genera un reporte PDF con los resultados de la verificacion."""

    COLOR_ECR_BLUE = colors.HexColor("#003366")
    COLOR_SUCCESS = colors.HexColor("#28a745")
    COLOR_ERROR = colors.HexColor("#dc3545")
    COLOR_WARNING = colors.HexColor("#ffc107")
    COLOR_GRAY_BG = colors.HexColor("#f8f9fa")

    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    fecha_hora = now.strftime("%Y-%m-%d %H:%M:%S")

    output_dir = Path("logs/reportes")
    output_dir.mkdir(parents=True, exist_ok=True)
    sufijo = "_post_correccion" if es_post_correccion else ""
    filename = output_dir / f"verificacion_grupos_{timestamp}{sufijo}.pdf"

    page_size = landscape(letter)
    doc = SimpleDocTemplate(
        str(filename), pagesize=page_size,
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
        topMargin=0.6 * inch, bottomMargin=0.7 * inch,
    )
    story = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='CustomTitle', parent=styles['Heading1'],
        fontSize=22, textColor=COLOR_ECR_BLUE, spaceAfter=12,
        alignment=TA_CENTER, fontName='Helvetica-Bold'
    ))
    styles.add(ParagraphStyle(
        name='CustomSubtitle', parent=styles['Heading2'],
        fontSize=14, textColor=COLOR_ECR_BLUE, spaceAfter=30,
        alignment=TA_CENTER, fontName='Helvetica'
    ))
    styles.add(ParagraphStyle(
        name='SectionHeader', parent=styles['Heading2'],
        fontSize=14, textColor=COLOR_ECR_BLUE, spaceAfter=12,
        fontName='Helvetica-Bold'
    ))
    # Estilo para celdas de tabla
    cell_style = ParagraphStyle(
        name='CellStyle', parent=styles['Normal'],
        fontSize=8, leading=10,
    )
    cell_header_white = ParagraphStyle(
        name='CellHeaderWhite', parent=styles['Normal'],
        fontSize=9, leading=11, fontName='Helvetica-Bold',
        textColor=colors.white,
    )
    cell_header_black = ParagraphStyle(
        name='CellHeaderBlack', parent=styles['Normal'],
        fontSize=9, leading=11, fontName='Helvetica-Bold',
        textColor=colors.black,
    )

    def P(text, style=cell_style):
        """Shortcut para crear Paragraph en celdas."""
        return Paragraph(str(text), style)

    # Estilo base para todas las tablas de datos
    BASE_TABLE_STYLE = [
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]

    # -- PORTADA --
    story.append(Spacer(1, 2 * inch))
    titulo = "Verificacion de Grupos de Licencia (Post-Corrección)" if es_post_correccion else "Verificacion de Grupos de Licencia"
    story.append(Paragraph(titulo, styles['CustomTitle']))
    story.append(Paragraph("Escuela Colombiana de Rehabilitacion", styles['CustomSubtitle']))
    story.append(Spacer(1, 0.5 * inch))

    info_data = [
        ["Fecha y hora de consulta:", fecha_hora],
        ["Archivo procesado:", "activos.xlsx"],
    ]
    info_table = Table(info_data, colWidths=[2.5 * inch, 4 * inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (0, -1), COLOR_ECR_BLUE),
    ]))
    story.append(info_table)
    story.append(PageBreak())

    # -- RESUMEN --
    story.append(Paragraph("Resumen", styles['SectionHeader']))
    story.append(Spacer(1, 0.2 * inch))

    summary_data = [
        ["Categoria", "Cantidad"],
        ["Total filas en Excel", str(total_filas)],
        ["Total procesados", str(total_procesados)],
        ["Omitidos (sin email/tipo)", str(len(omitidos))],
        ["Correctos (grupo correcto)", str(len(correctos))],
        ["En grupo incorrecto", str(len(incorrectos))],
        ["Sin grupo de licencia", str(len(sin_grupo))],
        ["Tipo de trabajador no reconocido", str(len(tipo_invalido))],
        ["Sobrantes (en grupo pero no en Excel)", str(len(sobrantes))],
    ]
    summary_table = Table(summary_data, colWidths=[4 * inch, 1.5 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), COLOR_ECR_BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, -1), COLOR_GRAY_BG),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.white),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)

    # -- USUARIOS EN GRUPO INCORRECTO --
    if incorrectos:
        story.append(PageBreak())
        story.append(Paragraph("Usuarios en Grupo Incorrecto", styles['SectionHeader']))
        story.append(Spacer(1, 0.2 * inch))

        table_data = [[
            P("Nombre", cell_header_white),
            P("Email", cell_header_white),
            P("Tipo", cell_header_white),
            P("Grupo Actual", cell_header_white),
            P("Grupo Esperado", cell_header_white),
        ]]
        for u in incorrectos:
            table_data.append([
                P(u['nombre']),
                P(u['email']),
                P(u['tipo']),
                P(", ".join(u['grupos_actuales'])),
                P(u['grupo_esperado']),
            ])

        t = Table(table_data, colWidths=[2.5 * inch, 2.5 * inch, 1.3 * inch, 2 * inch, 2 * inch])
        t.setStyle(TableStyle(BASE_TABLE_STYLE + [
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_ERROR),
        ]))
        story.append(t)

    # -- USUARIOS SIN GRUPO --
    if sin_grupo:
        story.append(PageBreak())
        story.append(Paragraph("Usuarios Sin Grupo de Licencia", styles['SectionHeader']))
        story.append(Spacer(1, 0.2 * inch))

        table_data = [[
            P("Nombre", cell_header_black),
            P("Email", cell_header_black),
            P("Tipo", cell_header_black),
            P("Deberia estar en", cell_header_black),
        ]]
        for u in sin_grupo:
            table_data.append([
                P(u['nombre']),
                P(u['email']),
                P(u['tipo']),
                P(u['grupo_esperado']),
            ])

        t = Table(table_data, colWidths=[2.8 * inch, 3 * inch, 1.5 * inch, 2.5 * inch])
        t.setStyle(TableStyle(BASE_TABLE_STYLE + [
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_WARNING),
        ]))
        story.append(t)

    # -- TIPO INVALIDO --
    if tipo_invalido:
        story.append(PageBreak())
        story.append(Paragraph("Usuarios con Tipo de Trabajador No Reconocido", styles['SectionHeader']))
        story.append(Spacer(1, 0.2 * inch))

        table_data = [[
            P("Nombre", cell_header_white),
            P("Email", cell_header_white),
            P("Tipo", cell_header_white),
        ]]
        for u in tipo_invalido:
            table_data.append([P(u['nombre']), P(u['email']), P(u['tipo'])])

        t = Table(table_data, colWidths=[3.5 * inch, 3.5 * inch, 2.5 * inch])
        t.setStyle(TableStyle(BASE_TABLE_STYLE + [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#6c757d")),
        ]))
        story.append(t)

    # -- SOBRANTES: EN GRUPO PERO NO EN EXCEL --
    if sobrantes:
        story.append(PageBreak())
        story.append(Paragraph(f"Sobrantes: En Grupo pero No en el Excel (se moverán a '{GRUPO_RETIRADOS}')", styles['SectionHeader']))
        story.append(Spacer(1, 0.2 * inch))

        table_data = [[
            P("Nombre", cell_header_white),
            P("Email", cell_header_white),
            P("Grupo", cell_header_white),
        ]]
        for u in sobrantes:
            table_data.append([
                P(u['nombre']),
                P(u['email']),
                P(u['grupo']),
            ])

        t = Table(table_data, colWidths=[3 * inch, 3.5 * inch, 3 * inch])
        t.setStyle(TableStyle(BASE_TABLE_STYLE + [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#6f42c1")),
        ]))
        story.append(t)

    # -- PIE DE PAGINA --
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.grey)
        canvas.drawString(
            inch, 0.4 * inch,
            f"Generado automaticamente - Verificacion de Grupos ECR - {fecha_hora}"
        )
        canvas.drawRightString(
            page_size[0] - inch, 0.4 * inch,
            f"Pagina {canvas.getPageNumber()}"
        )
        canvas.restoreState()

    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)

    logger.info(f"Reporte PDF generado: {filename}")
    return str(filename)


async def aplicar_cambios_interactivo(client, incorrectos: list, sin_grupo: list, sobrantes: list) -> bool:
    """
    Muestra los cambios pendientes y aplica los que el usuario confirme.
    Retorna True si se aplicó al menos una acción.
    """
    hay_cambios = incorrectos or sin_grupo or sobrantes
    if not hay_cambios:
        print("\nNo hay cambios pendientes.")
        return False

    se_aplico_algo = False

    # -- MOVER USUARIOS AL GRUPO CORRECTO --
    if incorrectos:
        print("\n" + "=" * 90)
        print("ACCION DISPONIBLE: Mover usuarios al grupo correcto")
        print("=" * 90)
        print(f"{'#':<4} {'NOMBRE':<40} {'EMAIL':<35} {'GRUPO ACTUAL':<32} {'GRUPO ESPERADO'}")
        print("-" * 90)
        for i, u in enumerate(incorrectos, 1):
            print(f"{i:<4} {u['nombre']:<40} {u['email']:<35} {', '.join(u['grupos_actuales']):<32} {u['grupo_esperado']}")

        respuesta = input("\n¿Deseas mover estos usuarios al grupo correcto? (S/N): ").strip().upper()
        if respuesta == "S":
            se_aplico_algo = True
            print("\nAplicando cambios...")
            ok = 0
            fail = 0
            for u in incorrectos:
                user_id = await client.get_user_id_by_email(u['email'])
                if not user_id:
                    print(f"  [ERROR] No se encontró el ID de {u['email']}")
                    fail += 1
                    continue

                # Remover solo de los grupos incorrectos (no del grupo esperado si ya está)
                for grupo_actual in u['grupos_actuales']:
                    if grupo_actual != u['grupo_esperado']:
                        grupo_actual_id = await client.get_group_id(grupo_actual)
                        if grupo_actual_id:
                            await client.remove_user_from_group(user_id, grupo_actual_id, grupo_actual)

                # Agregar al grupo correcto (solo si aún no está)
                grupo_esperado_id = await client.get_group_id(u['grupo_esperado'])
                if grupo_esperado_id:
                    exito = await client.add_user_to_group(user_id, grupo_esperado_id, u['grupo_esperado'])
                    if exito:
                        print(f"  [OK] {u['nombre']} -> {u['grupo_esperado']}")
                        ok += 1
                    else:
                        print(f"  [ERROR] No se pudo agregar {u['email']} a {u['grupo_esperado']}")
                        fail += 1
                else:
                    print(f"  [ERROR] No se encontró el grupo: {u['grupo_esperado']}")
                    fail += 1

            print(f"\nResultado: {ok} movidos correctamente, {fail} errores.")
        else:
            print("Operacion cancelada.")

    # -- AGREGAR USUARIOS SIN GRUPO AL GRUPO CORRECTO --
    if sin_grupo:
        print("\n" + "=" * 90)
        print("ACCION DISPONIBLE: Agregar usuarios sin grupo al grupo que les corresponde")
        print("=" * 90)
        print(f"{'#':<4} {'NOMBRE':<40} {'EMAIL':<35} {'DEBERIA ESTAR EN'}")
        print("-" * 90)
        for i, u in enumerate(sin_grupo, 1):
            print(f"{i:<4} {u['nombre']:<40} {u['email']:<35} {u['grupo_esperado']}")

        respuesta = input("\n¿Deseas agregar estos usuarios a su grupo correspondiente? (S/N): ").strip().upper()
        if respuesta == "S":
            se_aplico_algo = True
            print("\nAplicando cambios...")
            ok = 0
            fail = 0
            for u in sin_grupo:
                user_id = await client.get_user_id_by_email(u['email'])
                if not user_id:
                    print(f"  [ERROR] No se encontró el ID de {u['email']}")
                    fail += 1
                    continue

                grupo_id = await client.get_group_id(u['grupo_esperado'])
                if not grupo_id:
                    print(f"  [ERROR] No se encontró el grupo: {u['grupo_esperado']}")
                    fail += 1
                    continue

                exito = await client.add_user_to_group(user_id, grupo_id, u['grupo_esperado'])
                if exito:
                    print(f"  [OK] {u['nombre']} -> {u['grupo_esperado']}")
                    ok += 1
                else:
                    print(f"  [ERROR] No se pudo agregar {u['email']} a {u['grupo_esperado']}")
                    fail += 1

            print(f"\nResultado: {ok} agregados correctamente, {fail} errores.")
        else:
            print("Operacion cancelada.")

    # -- MOVER SOBRANTES A GRUPO RETIRADOS --
    if sobrantes:
        print("\n" + "=" * 90)
        print(f"ACCION DISPONIBLE: Mover sobrantes a '{GRUPO_RETIRADOS}'")
        print("=" * 90)
        print(f"{'#':<4} {'NOMBRE':<40} {'EMAIL':<35} {'GRUPO ACTUAL'}")
        print("-" * 90)
        for i, u in enumerate(sobrantes, 1):
            print(f"{i:<4} {u['nombre']:<40} {u['email']:<35} {u['grupo']}")

        respuesta = input(f"\n¿Deseas mover estos usuarios a '{GRUPO_RETIRADOS}'? (S/N): ").strip().upper()
        if respuesta == "S":
            se_aplico_algo = True
            print("\nAplicando cambios...")
            ok = 0
            fail = 0

            grupo_retirados_id = await client.get_group_id(GRUPO_RETIRADOS)
            if not grupo_retirados_id:
                print(f"  [ERROR FATAL] No se encontró el grupo destino: '{GRUPO_RETIRADOS}'")
                return se_aplico_algo

            for u in sobrantes:
                user_id = await client.get_user_id_by_email(u['email'])
                if not user_id:
                    print(f"  [ERROR] No se encontró el ID de {u['email']}")
                    fail += 1
                    continue

                # Remover del grupo actual (A5)
                grupo_origen_id = await client.get_group_id(u['grupo'])
                if not grupo_origen_id:
                    print(f"  [ERROR] No se encontró el grupo origen: {u['grupo']}")
                    fail += 1
                    continue

                removido = await client.remove_user_from_group(user_id, grupo_origen_id, u['grupo'])
                if not removido:
                    print(f"  [ERROR] No se pudo remover {u['email']} de {u['grupo']}")
                    fail += 1
                    continue

                # Agregar al grupo de retirados
                agregado = await client.add_user_to_group(user_id, grupo_retirados_id, GRUPO_RETIRADOS)
                if agregado:
                    print(f"  [OK] {u['nombre']} -> {GRUPO_RETIRADOS}")
                    ok += 1
                else:
                    print(f"  [ERROR] Se removió de '{u['grupo']}' pero no se pudo agregar a '{GRUPO_RETIRADOS}'")
                    fail += 1

            print(f"\nResultado: {ok} movidos a retirados correctamente, {fail} errores.")
        else:
            print("Operacion cancelada.")

    return se_aplico_algo


if __name__ == "__main__":
    logger.add(
        "logs/verify_groups.log",
        rotation="10 MB",
        level="DEBUG",
        format="{time:YYYY-MM-DD HH:mm:ss} | {level: <8} | {message}"
    )
    asyncio.run(check_group_assignments())
